import pandas as pd
import numpy as np
import os
from html2excel import ExcelParser
from openpyxl import load_workbook
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

def adjust_column_width(file_path):
    """Automatically adjust Excel column widths to fit content.
    
    Args:
        file_path: Path to Excel file to adjust
    """
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0  
            from openpyxl.utils import get_column_letter
            # Find the first cell in the column that is not a merged cell and has a valid column index
            first_cell = next((cell for cell in col if hasattr(cell, "column") and cell.column is not None), None)
            if first_cell is None or first_cell.column is None:
                continue  # Skip columns where column index is None
            column = get_column_letter(first_cell.column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
    wb.save(file_path)
    wb.close()

def html_to_excel(html_file, excel_file):
    """Convert HTML file to Excel format.
    
    Args:
        html_file: Path to input HTML file
        excel_file: Path to output Excel file
    """
    tables = pd.read_html(html_file, encoding='latin1', decimal=',', thousands='.')
    df = tables[0]
    df.to_excel(excel_file, index=False, engine='openpyxl')

def process_isso(file_name, base_path=os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))):
    """Process ISSO energy consumption data files.
    
    This function handles complex energy consumption reports from the ISSO platform:
    1. Converts .xls files to proper XLSX
    2. Processes multi-parameter electrical measurements
    3. Adds UTC timezone calculations
    4. Calculates power consumption in kW and energy in kWh
    5. Handles extensive electrical parameter datasets
    6. Updates database with processed consumption data
    
    Args:
        file_name: Name of the raw data file to process
        base_path: Base directory path for file operations
    """
    raw_folder = os.path.join(base_path, 'downloads', 'dados_brutos', 'isso')
    treated_excel_folder = os.path.join(base_path, 'downloads', 'dados_tratados', 'isso', 'planilha')
    treated_csv_folder = os.path.join(base_path, 'downloads', 'dados_tratados', 'isso', 'csv')
    
    file_path = os.path.join(raw_folder, file_name)
    base_name = os.path.splitext(file_name)[0]

    # Paths for saving treated file
    new_file_path = os.path.join(treated_excel_folder, 'PROCESSED_' + base_name + ".xlsx")
    file_path_csv = os.path.join(treated_csv_folder, 'PROCESSED_' + base_name + ".csv")

    translated_columns = ["UTC-3 Start","UTC-3 End","Frequency","Voltage A","Voltage B","Voltage C","Voltage A-B","Voltage B-C","Voltage C-A","Current A","Current B","Current C","Neutral Current Measured","Neutral Current Calculated","Apparent Power A","Apparent Power B","Apparent Power C","Apparent Power Arithmetic Sum","Apparent Power Vector Sum","Reactive Power A","Reactive Power B","Reactive Power C","Total Reactive Power","Real Power Factor A","Real Power Factor B","Real Power Factor C","Average Real Power Factor","Cos(φ) A","Cos(φ) B","Cos(φ) C","Average Cos(φ)","Phase Inductive 1","Phase Inductive 2","Phase Inductive 3","Average Inductive","Direct: Active Power Fund+Harm A","Direct: Active Power Fund+Harm B","Direct: Active Power Fund+Harm C","Direct: Total Active Power Fund+Harm","Direct: Active Power Fundamental A","Direct: Active Power Fundamental B","Direct: Active Power Fundamental C","Direct: Total Active Power Fundamental","Direct: Active Power Harmonic A","Direct: Active Power Harmonic B","Direct: Active Power Harmonic C","Direct: Total Active Power Harmonic","Reverse: Active Power Fund+Harm A","Reverse: Active Power Fund+Harm B","Reverse: Active Power Fund+Harm C","Reverse: Total Active Power Fund+Harm","Reverse: Active Power Fundamental A","Reverse: Active Power Fundamental B","Reverse: Active Power Fundamental C","Reverse: Total Active Power Fundamental","Reverse: Active Power Harmonic A","Reverse: Active Power Harmonic B","Reverse: Active Power Harmonic C","Reverse: Total Active Power Harmonic","Voltage Unbalance (Phasor) Total","Voltage Unbalance (Amplitude) Total","Current Unbalance (Amplitude) Total"] 

    html = False
    # Convert .xls to .xlsx
    try:
        xls = pd.ExcelFile(file_path, engine='xlrd')
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=3, decimal=',')
                df.columns = translated_columns
                df.to_excel(writer, sheet_name=str(sheet_name), index=False)
    except Exception as e:
        html = True
        html_to_excel(file_path, new_file_path)
        df = pd.read_excel(new_file_path, engine='openpyxl', skiprows=4, decimal=',')
        df.columns = translated_columns
      
    # Read converted file with Pandas

    df['UTC-3 Start'] = pd.to_datetime(df['UTC-3 Start'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
    df['UTC-3 End'] = pd.to_datetime(df['UTC-3 End'], format='%d/%m/%Y %H:%M:%S', errors='coerce')

    # Add UTC columns
    df['UTC Start'] = df['UTC-3 Start'] - pd.Timedelta(hours=-3)
    df['UTC End'] = df['UTC-3 End'] - pd.Timedelta(hours=-3)

    # Reorder columns to place UTC columns after the first two
    cols = list(df.columns)
    cols.remove('UTC Start')
    cols.remove('UTC End')
    new_order = cols[:2] + ['UTC Start', 'UTC End'] + cols[2:]
    df = df[new_order]

    # Create kW demand columns (each value divided by 1000)

    df['Direct: Total Active Power Fund+Harm (kW)'] = df['Direct: Total Active Power Fund+Harm'] / 1000
    df['Reverse: Total Active Power Fund+Harm (kW)'] = df['Reverse: Total Active Power Fund+Harm'] / 1000

    # Calculate kWh totals (each value divided by 1000 * 0.25)
    df['Direct: Total kWh'] = df['Direct: Total Active Power Fund+Harm'] / 1000 * 0.25
    df['Reverse: Total kWh'] = df['Reverse: Total Active Power Fund+Harm'] / 1000 * 0.25
    df.to_csv(file_path_csv, index=False, encoding='utf-8')
    df.to_excel(new_file_path, index=False, engine='openpyxl')
    adjust_column_width(new_file_path)